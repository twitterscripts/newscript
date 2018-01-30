# -*- coding: utf-8 -*-
import os
import json
import oauth2 as oauth
import openpyxl
import time
import tweepy
import operator
import openpyxl
from openpyxl import Workbook
start_time = time.clock()

#############################
## PROGRAM DESCRIPTION ######
#############################

"""
The purpose of this program is to output the following information:

- In a sector (on Twitter), who are the key influencers?
- Which influencers are most influential for each brand?
- Which influencers are most influential for each brand's competitors?
(can also be extended to...)
- Which non-sector brands in a sector are ideal for partnership for the selected brand?
- Which news organizations or media channels are best for optimizing advertising on a brand-specific basis?

This program TAKES IN:

A list of key competitor twitter handles in a sector. For example, in the French 'sports wear'
sector, this would be a list like: ["adidasFR","ReebokFR","NikeFrance","UnderArmourFR","NewBalanceFR","ASICSFrance","lecoqsportif"]

This program OUTPUTS:

An excel file which has the following information:

- One "Sector" sheet that ranks key influencers (including brands, news organizations, etc) in the sector based on the
preferences of the followers of all of the listed competitive brands
- One sheet per listed competitor which summarizes the influencer counts for that selected brand's followers. This sheet also contains
the influencers who are followed by the followers of their competitors but not that brand itself. For example, if I'm
on the "Nike" sheet, I can see Nike's followers' influencers and also the combined and ranked influencers of all other
competitors (Adidas, Reebok, Puma, etc) who do not appear in Nike's list. This would allow Nike to hire influencers
who might previously not have been on their radar to reach their competitors' followers.

METHOD:

- This program collects the last N Twitter followers for each brand in an inputted list.
- It finds the profiles that these N followers of that brand follow in common and ranks them by occurence count.
- It combines and sorts this data and then outputs an excel file.

TO CALCULATE POTENTIAL RUNTIMES:
(in minutes)

MINIMUM runtime (very unlikely) = "cutoff" value from below X number of brands included in 'list_of_competitors'
MAXIMUM runtime (likely) = 3 X "cutoff" value from below X number of brands included in 'list_of_competitors'

So if you're looking at 100 Follower IDs for a list of 3 brands each, at minimum it will 300 minutes, and at maximum it will take around 900 minutes.

"""

#############################################################
###### ADJUSTABLE VALUES: MAKE YOUR CHANGES HERE! ###########
#############################################################

#This is the number of followers you want to take and process influencers for per listed competitor.
#The higher the number, the longer the program takes to run.
cutoff = 5

### Update this list with the twitter handles of the competing brands in your chosen sector. The more competitors you include, the longer the program takes to run.
#EXAMPLE: list_of_competitors = ["adidasFR","ReebokFR","NikeFrance","UnderArmourFR","NewBalanceFR","ASICSFrance","lecoqsportif"]
list_of_competitors = ["pizzahut","dominos","PapaJohns"]

### Put the path you will save the data in
path = "/Users/softops/Desktop/Django/operationcc/serra/"

### HERE, PUT THE GENERAL NAME OF THE SECTOR YOU ARE EXPLORING. Always give this section a different name or it will
### overwrite previously created folders and files. Even if you are doing multiple batches of sports accounts, give them
### sector names like SPORTS_1, SPORTS_2, etc to keep them distinct.
sector = "PIZZA"

#TWITTER CREDS
#This you get from Twitter's website with your personal handle. If you're not sure about this one, you can ask for
#help from a developer.

consumer_key = "Gw9Xnme7QqJ7jwWNx3qIGIbn7"
consumer_secret = "uZeReNwSMNB2B8VzlVizhoAB05zevvrWdtsMSiBisVxI9ImDgC"
access_token = "929636920619126784-eYZxALHZBRUhq23thpmqN8JqRcneBWP"
access_token_secret ="h4yh9ilDvwbMsDYGEEYSGarVRXFwWo9SAxb9My2IyhRQX"

#############################################################
###### BELOW HERE, NO TOUCHY! ###############################
#############################################################

auth = tweepy.OAuthHandler(consumer_key,consumer_secret)
auth.set_access_token(access_token, access_token_secret)
api = tweepy.API(auth)

dict_file_names = {}
dict_influencers_master = {}
dict_descriptions_master = {}
dict_follower_counts_master = {}

path_addition = "Collected_Brand_Followers_" + sector
combined_path = path + path_addition

#Initialize Folder in which follower data for each brand will be collected.
def initialize():
    try:
        os.mkdir(combined_path)
    except:
        return

#Collect last N followers for a given screen_name and save as text file
def collect_N_followers(screen_name):
    text_file_name = combined_path +"/"+ screen_name + "Followers_IDs.txt"
    text_file = open(text_file_name, "w") #open text file to store follower IDs
    dict_file_names[screen_name] = text_file_name #keep track of file locations
    counter = 0
    for page in tweepy.Cursor(api.followers_ids, screen_name=screen_name).pages():
        for num in page:
            text_file.write(str(num) + "\n") #store each ID
            counter = counter + 1
            if counter == cutoff: #If we reach the "cutoff" number of IDs to process for each brand, close the text file and leave this function. You can set this manually as well.
                text_file.close()
                print (screen_name + "file saved.")
                return

# Cycles through listed competitors and collects and saves followers for them
def collect_all_competitors():
    for i in list_of_competitors:
        print ("Collecting followers for: "+i)
        collect_N_followers(i)


#Reads in and returns a list of follower IDs for a brand from a saved file
def return_list_of_ids(brand_key):
    return_list = []
    readfile = open(dict_file_names[brand_key],"r") #look up where the saved file is from dict_file_names
    counter = 0
    for i in readfile.readlines():
        counter = counter + 1
        entry = i.replace("\n","")
        return_list = return_list+[entry]
        #if counter == cutoff: #Only to be unchecked if you want to save more IDs than you intend to get influencers for
         #   break
    return return_list


### For a twitter ID, returns list of Twitter accounts this ID is following (the "influencers")
def get_friends(twitterid):
    print ("Now processing Twitter ID: "+str(twitterid))
    friends = []
    count = 0
    for friend in tweepy.Cursor(api.friends, id=twitterid, wait_on_rate_limit=True, count=200).pages():
        #print len(friend)
        friends.extend(friend)
        count = count+1
        time.sleep(61) #to prevent the program from getting stuck on the rate limit
        if count == 3:  #No more than 600 "following." You can increase this if you want to get more "following" accounts for each ID.
            break
    m = []
    for entry in friends:
        entryname = entry.screen_name
        m = m+[entryname]

        entry_descr = entry.description
        entry_follower_count = entry.followers_count
        dict_descriptions_master[entryname] = entry_descr
        dict_follower_counts_master[entryname] = entry_follower_count

    print ("Completed processing Twitter ID: "+str(twitterid))
    return m


#Updates master dictionary with influencer counts for each brand
#in the format: BRAND: Dictionary of influencers (Influencer: count)
# (One way to make sure this function works is that the total count for the brand itself within dict_influencers_temp should match the number of IDs you process since they're all followers of that brand)

def update_influencers_dict(brandx):
    dict_influencers_temp = {} #this serves as the dictionary to collect total influencer counts for a single brand
    selected_followers = return_list_of_ids(brandx) #Return follower IDs from a saved file
    for profile in selected_followers: #for each ID
        try:
            list_of_influencers = get_friends(profile) #Gets the profiles this ID is following
            for id in list_of_influencers:
                if id not in dict_influencers_temp: #Counts the number of times each influencer occurs for this brand (occurence among the followers "friends")
                    dict_influencers_temp[id] = 1
                else:
                    dict_influencers_temp[id] = dict_influencers_temp[id]+1
        except:
            print ("Could not process the Twitter Profile: "+str(profile))
            pass

    dict_influencers_master[brandx] = dict_influencers_temp #Saves brand influencer dictionary under brand name within master dictionary

#Cycles through brands and updates master dictionary
def process_all_brands():
    for brand_name in list_of_competitors:
        print ("Updating Master Dictionary with influencers for "+brand_name)
        update_influencers_dict(brand_name)
        print (brand_name+" Master Dictionary influencer update completed.")

#Collects counts from all brands within an inputted dictionary
def return_total_counts(selected_brands_dict):
    temp_totals_dict = {} #dictionary to collect total counts
    for key in selected_brands_dict: #for each brand
        temp_dict = dict_influencers_master[key] #dictionary of brand-specific influencers for that particular brand that was saved under that brand key
        for subkey in temp_dict: #for each influencer for that brand
            if subkey not in temp_totals_dict:
                temp_totals_dict[subkey] = temp_dict[subkey] #write the matching value from the first brand in which you encounter that influencer
            else:
                temp_totals_dict[subkey] = temp_totals_dict[subkey] + temp_dict[subkey]
    return temp_totals_dict

#removes key from dictionary and returns a copy
def removekey(d, key):
    r = dict(d)
    del r[key]
    return r

#Return dictionaries containing counts for each of the "brand" sheets
def return_brand_counts(brandx):
    brand_dict_t = dict_influencers_master[brandx] #chosen brand dictionary
    competitors_dict = removekey(dict_influencers_master,brandx) #returns a copy of the master dictionary with our chosen brand removed
    influencers_competitors = return_total_counts(competitors_dict)
    comp_totals_dict = {}


    for key in influencers_competitors: #for each influencer in our competitors' collected dictionary
        if key not in brand_dict_t: #if this influencer doesn't appear in our chosen brand's influencers
            comp_totals_dict[key] = influencers_competitors[key] #copy it to the dictionary we will return

    return brand_dict_t,comp_totals_dict

#Initializes Excel sheet and calls necessary functions to complete each stage of the write-out.
def initialize_spreadsheets():
    print ("Initializing Spreadsheet.")
    wl = Workbook() # create new workbook for output

    #"SECTOR" Sheet with totals
    print ("Writing Sector Page.")
    master_sheet = wl.create_sheet('Sector')
    master_sheet["A1"] = "Influencer Name"
    master_sheet["B1"] = "Total Count"
    master_sheet["C1"] = "Follower Count"
    master_sheet["D1"] = "Twitter Description"
    master_dict = return_total_counts(dict_influencers_master) #Call function to process total results with dictionary containing all brands
    master_counter = 2

    sorted_master = sorted(master_dict.items(), key=operator.itemgetter(1), reverse=True)

    for key in sorted_master: #for each tuple in the sorted dictionary
        influencer_name = key[0]
        influencer_count = key[1]
        master_sheet["A" + str(master_counter)] = influencer_name #influencer name
        master_sheet["B"+str(master_counter)] = influencer_count #influencer count
        master_sheet["C" + str(master_counter)] = dict_follower_counts_master[influencer_name]
        try:
            master_sheet["D" + str(master_counter)] = dict_descriptions_master[influencer_name]
        except:
            pass
        master_counter = master_counter + 1

    print ("Sector Page complete.")

    #Brand-specific Sheets
    for brandx in list_of_competitors:
        print ("Writing page for "+brandx)
        temp_sheet = wl.create_sheet(brandx)
        brand_dict_m, competitors_dict_m = return_brand_counts(brandx)

        # Write out influencers for a brand
        temp_sheet["A1"] = brandx+" Influencers"
        temp_sheet["B1"] = "Count"
        temp_sheet["C1"] = "Follower Count"
        temp_sheet["D1"] = "Twitter Description"
        brand_counter = 2

        sorted_brand = sorted(brand_dict_m.items(), key=operator.itemgetter(1), reverse=True)

        for key in sorted_brand:
            influencer_name = key[0]
            influencer_count = key[1]
            temp_sheet["A"+str(brand_counter)] = influencer_name
            temp_sheet["B"+str(brand_counter)] = influencer_count
            temp_sheet["C" + str(brand_counter)] = dict_follower_counts_master[influencer_name]
            try:
                temp_sheet["D" + str(brand_counter)] = dict_descriptions_master[influencer_name]
            except:
                pass
            brand_counter = brand_counter + 1

        # Write out influencers for the brand's competitors
        temp_sheet["F1"] = "Competitor Influencers"
        temp_sheet["G1"] = "Count"
        temp_sheet["H1"] = "Follower Count"
        temp_sheet["I1"] = "Twitter Description"
        competitor_counter = 2

        sorted_competitors = sorted(competitors_dict_m.items(), key=operator.itemgetter(1), reverse=True)

        for key in sorted_competitors:
            influencer_name = key[0]
            influencer_count = key[1]
            temp_sheet["F"+str(competitor_counter)] = influencer_name
            temp_sheet["G"+str(competitor_counter)] = influencer_count
            temp_sheet["H" + str(competitor_counter)] = dict_follower_counts_master[influencer_name]
            try:
                temp_sheet["I" + str(competitor_counter)] = dict_descriptions_master[influencer_name]
            except:
                pass
            competitor_counter = competitor_counter + 1

        print (brandx + " Page complete.")

    #remove default sheet
    std = wl.get_sheet_by_name('Sheet')
    wl.remove_sheet(std)

    print ("Excel write out complete.")
    wl.save(path+"Influencers_"+sector+".xlsx")
    print ("File saved as: Influencers_"+sector+".xlsx")

#the part of the program that runs everything
def main():
    initialize()
    collect_all_competitors()
    process_all_brands()
    initialize_spreadsheets()
    print ("Program Complete.")
    print (time.clock() - start_time, "seconds")

main()